import requests
from datetime import datetime ,timedelta
import pandas as pd
import time
from openpyxl import load_workbook

Bot_Token = "5260912354:AAGyOqFjyo38L9HZe1IBSxrCR05DCsNxOdg"

Grade_1_Group_Id = "-1001789675957" #1001789675957
Grade_2_Group_Id = "-1001506290416"
Grade_3_Group_Id = "-1001496283204"
Grade_4_Group_Id = "-1001196265000"

LastDate_Sent = "09,12,2022"
Message_Sent_Status = False
Grade_1_Path = "Grade_1/"
Grade_2_Path = "Grade_2/"
Grade_3_Path = "Grade_3/"
Grade_4_Path = "Grade_4/"

def Getnowdate():
    return datetime.now().strftime(f"%d,%m,%Y")

def SaveMessageId( MessageId , SheetIndex , DayLineNum ):
    book = load_workbook("Schedule_Info.xlsx")
    sheet = book.worksheets[SheetIndex]
    sheet[f"D{DayLineNum}"] = MessageId
    book.save("Schedule_Info.xlsx")
    book.save("Schedule_Info - Copy.xlsx")
    book.close()

def SendMessage(Message_Param , SheetIndex , LastMessageId , DayLineNum  ):
    try:

        global Message_Sent_Status
        global LastDate_Sent

        if (SheetIndex == 0):
            Files = {'photo':open(f'{Grade_1_Path + Message_Param}',"rb")}
            RespnseSend = requests.post(f"https://api.telegram.org/bot{Bot_Token}/sendphoto?chat_id={Grade_1_Group_Id}&caption=الورد اليومي للقران الكريم",files=Files)
            RespnseDelete = requests.post(f"https://api.telegram.org/bot{Bot_Token}/deletemessage?chat_id={Grade_1_Group_Id}&message_id={LastMessageId}")
            print(f"Quran Page Of Date : {Getnowdate()} Has Been Sent And Response Status Is : {RespnseSend.status_code}")
            print(f"Quran Page Of Date : {LastDate_Sent} Has Been Deleted And Response Status Is : {RespnseDelete.status_code}")
            MessageId = (RespnseSend.json()["result"]["message_id"])
            SaveMessageId(MessageId , SheetIndex , DayLineNum  )
        elif (SheetIndex == 1):
            Files = {'photo':open(f'{Grade_2_Path + Message_Param}',"rb")}
            RespnseSend = requests.post(f"https://api.telegram.org/bot{Bot_Token}/sendphoto?chat_id={Grade_2_Group_Id}&caption=الورد اليومي للقران الكريم",files=Files)
            RespnseDelete = requests.post(f"https://api.telegram.org/bot{Bot_Token}/deletemessage?chat_id={Grade_2_Group_Id}&message_id={LastMessageId}")
            print(f"Quran Page Of Date : {Getnowdate()} Has Been Sent And Response Status Is : {RespnseSend.status_code}")
            print(f"Quran Page Of Date : {LastDate_Sent} Has Been Deleted And Response Status Is : {RespnseDelete.status_code}")
            MessageId = (RespnseSend.json()["result"]["message_id"])
            SaveMessageId(MessageId , SheetIndex , DayLineNum  )
        elif (SheetIndex == 2):
            Files = {'photo':open(f'{Grade_3_Path + Message_Param}',"rb")}
            RespnseSend = requests.post(f"https://api.telegram.org/bot{Bot_Token}/sendphoto?chat_id={Grade_3_Group_Id}&caption=الورد اليومي للقران الكريم",files=Files)
            RespnseDelete = requests.post(f"https://api.telegram.org/bot{Bot_Token}/deletemessage?chat_id={Grade_3_Group_Id}&message_id={LastMessageId}")
            print(f"Quran Page Of Date : {Getnowdate()} Has Been Sent And Response Status Is : {RespnseSend.status_code}")
            print(f"Quran Page Of Date : {LastDate_Sent} Has Been Deleted And Response Status Is : {RespnseDelete.status_code}")
            MessageId = (RespnseSend.json()["result"]["message_id"])
            SaveMessageId(MessageId , SheetIndex , DayLineNum  )
        elif (SheetIndex == 3):
            Files = {'photo':open(f'{Grade_4_Path + Message_Param}',"rb")}
            RespnseSend = requests.post(f"https://api.telegram.org/bot{Bot_Token}/sendphoto?chat_id={Grade_4_Group_Id}&caption=الورد اليومي للقران الكريم",files=Files)
            RespnseDelete = requests.post(f"https://api.telegram.org/bot{Bot_Token}/deletemessage?chat_id={Grade_4_Group_Id}&message_id={LastMessageId}")
            print(f"Quran Page Of Date : {Getnowdate()} Has Been Sent And Response Status Is : {RespnseSend.status_code}")
            print(f"Quran Page Of Date : {LastDate_Sent} Has Been Deleted And Response Status Is : {RespnseDelete.status_code}")
            MessageId = (RespnseSend.json()["result"]["message_id"])
            SaveMessageId(MessageId , SheetIndex , DayLineNum  )

        Message_Sent_Status = True
        LastDate_Sent = Getnowdate()

    except:
        pass

def Open_Schedule_Info_Excel( SheetName , SheetIndex ):
    try:  
            global Message_Sent_Status
            global LastDate_Sent

            Schedule_Info_Excel = pd.read_excel( io ="Schedule_Info.xlsx" , sheet_name = SheetName )
            Dates_List = list(Schedule_Info_Excel["Date"])
            Day_Status_List = list(Schedule_Info_Excel["Status"])
            Day_Line_Num = Dates_List.index(Getnowdate()) + 2

            
            book = load_workbook("Schedule_Info.xlsx")
            sheet = book.worksheets[SheetIndex]
            sheet[f"C{Day_Line_Num}"] = 1
            Message = sheet[f"B{Day_Line_Num}"].value
            Last_Message_Id = sheet[f"D{Day_Line_Num - 1}"].value
            book.save("Schedule_Info.xlsx")
            book.save("Schedule_Info - Copy.xlsx")
            book.close()

            if (Day_Status_List[ Day_Line_Num -2 ] == 0 ):
                    SendMessage(Message , SheetIndex , Last_Message_Id , Day_Line_Num  )
            else:
                Message_Sent_Status = True
                LastDate_Sent = Getnowdate()
    except:
        pass

while True:
    try:
            
        if ( (LastDate_Sent != Getnowdate()) and (Message_Sent_Status == True) ) or ( (LastDate_Sent != Getnowdate()) and (Message_Sent_Status == False ) ) or ( (LastDate_Sent == Getnowdate()) and (Message_Sent_Status == False ) ) :   
            Open_Schedule_Info_Excel(SheetName = "Grade_1" , SheetIndex=  0 )    
            Open_Schedule_Info_Excel(SheetName = "Grade_2" , SheetIndex=  1 )  
            Open_Schedule_Info_Excel(SheetName = "Grade_3" , SheetIndex=  2 )   
            Open_Schedule_Info_Excel(SheetName = "Grade_4" , SheetIndex=  3 )  
            
        #time.sleep(21600)

    except:
        pass